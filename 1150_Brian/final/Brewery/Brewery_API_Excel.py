"""
Mark Porraz
8/6/2024
Brewery_API.py

Here is where I make a call to pull 20 randomly selected breweries and place them in an Excel document

Here is the grand website where data is pulled from:
https://www.openbrewerydb.org/documentation#by_dist

"""

import requests
import random
from openpyxl import Workbook


# step 1 make API call
url = 'https://api.openbrewerydb.org/v1/breweries'
response_list_of_all_breweries = requests.get(url).json()
# print(brewery_list_url), no longer needed --- used for checking


# step 2: select from a list of 20 randomly selected breweries
list_of_random_breweries_chosen = random.sample(response_list_of_all_breweries, 20)
# print(list_of_random_breweries_chosen) --- used for checking

# Step 3: Create an Excel workbook and worksheet
brewery_wb = Workbook()
worksheet = brewery_wb.active
worksheet.title = 'Brewery Guide'

# Step 4: Add headers to the worksheet (move this outside the loop)
headers = ['Brewer Name', 'Address', 'City', 'State', 'Brewery Type', 'Website URL']

for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num, value=header)

# Initialize the row counter for the data
row = 2

# Step 5: Iterate over the selected breweries and get details
for one_brewery in list_of_random_breweries_chosen:
    brewery_id = one_brewery['id']
    url_for_one_brewery_details = f'https://api.openbrewerydb.org/v1/breweries/{brewery_id}'

    one_brewery_details_response = requests.get(url_for_one_brewery_details)

    if one_brewery_details_response.status_code == 200:  # If request is successful
        one_brewery_details = one_brewery_details_response.json()

        # Extracting details from the brewery
        brewery_name = one_brewery_details.get('name')
        brewery_address = one_brewery_details.get('address_1')
        brewery_city = one_brewery_details.get('city')
        brewery_state = one_brewery_details.get('state')
        brewery_type = one_brewery_details.get('brewery_type')
        brewery_website = one_brewery_details.get('website_url')

        # Adding the details to the worksheet
        worksheet.cell(row=row, column=1, value=brewery_name)
        worksheet.cell(row=row, column=2, value=brewery_address)
        worksheet.cell(row=row, column=3, value=brewery_city)
        worksheet.cell(row=row, column=4, value=brewery_state)
        worksheet.cell(row=row, column=5, value=brewery_type)
        worksheet.cell(row=row, column=6, value=brewery_website)

        row += 1
    else:
        print(f"Failed to retrieve details for brewery ID: {brewery_id}")

# # step 6
brewery_wb.save('Brewery_Guide.xlsx')




