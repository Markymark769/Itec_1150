"""
Mark Porraz
8/6/2024
Brewery_API.py

Here is where I make a call to pull 20 randomly selected breweries and place them in an Excel document

Here is the grand website where data is pulled from:
https://www.openbrewerydb.org/documentation#by_dist
"""

import requests
import random
from openpyxl import Workbook

# Step 1: Make the API call
url = 'https://api.openbrewerydb.org/v1/breweries'
response = requests.get(url)

# Check if the request was successful
if response.status_code == 200:
    response_list_of_all_breweries = response.json()
else:
    print("Failed to retrieve data")
    response_list_of_all_breweries = []

# Step 2: Select 20 random breweries from the list
if len(response_list_of_all_breweries) >= 20:
    list_of_random_breweries_chosen = random.sample(response_list_of_all_breweries, 20)
else:
    list_of_random_breweries_chosen = response_list_of_all_breweries

# Step 3: Create an Excel workbook and worksheet
brewery_wb = Workbook()
worksheet = brewery_wb.active
worksheet.title = 'Brewery Guide'

# Step 4: Add headers to the worksheet (move this outside the loop)
headers = ['Brewer Name', 'Address', 'City', 'State', 'Brewery Type', 'Website URL']
for col_num, header in enumerate(headers, 1):
    worksheet.cell(row=1, column=col_num, value=header)

# Initialize the row counter for the data
row = 2

# Step 5: Iterate over the selected breweries and get details
for one_brewery in list_of_random_breweries_chosen:
    brewery_id = one_brewery['id']
    url_for_one_brewery_details = f'https://api.openbrewerydb.org/v1/breweries/{brewery_id}'

    one_brewery_details_response = requests.get(url_for_one_brewery_details)

    if one_brewery_details_response.status_code == 200:  # If request is successful
        one_brewery_details = one_brewery_details_response.json()

        # Extracting details from the brewery
        brewery_name = one_brewery_details.get('name', 'N/A')
        brewery_address = one_brewery_details.get('address_1', 'N/A')
        brewery_city = one_brewery_details.get('city', 'N/A')
        brewery_state = one_brewery_details.get('state', 'N/A')
        brewery_type = one_brewery_details.get('brewery_type', 'N/A')
        brewery_website = one_brewery_details.get('website_url', 'N/A')

        # Adding the details to the worksheet
        worksheet.cell(row=row, column=1, value=brewery_name)
        worksheet.cell(row=row, column=2, value=brewery_address)
        worksheet.cell(row=row, column=3, value=brewery_city)
        worksheet.cell(row=row, column=4, value=brewery_state)
        worksheet.cell(row=row, column=5, value=brewery_type)
        worksheet.cell(row=row, column=6, value=brewery_website)

        row += 1
    else:
        print(f"Failed to retrieve details for brewery ID: {brewery_id}")

# Step 6: Save the Excel document
brewery_wb.save('Brewery_Guide.xlsx')
