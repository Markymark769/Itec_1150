from bs4 import BeautifulSoup
import requests
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

filename = 'ITEC_Schedules.xlsx'
wb = Workbook()  # make a new workbook
dt = datetime.datetime.now()  # get the current time
month = dt.month  # determine current month
ay = {1: 'Summer', 3: 'Fall', 5: 'Spring'}  # setup academic year periods

if 6 <= month <= 7:  # see if we're in the summer
    yrtrs = [str(dt.year + 1) + str(1), str(dt.year + 1) + str(3)]  # of the next year
    semesters = [f'{ay[1]} {str(dt.year)}', f'{ay[3]} {str(dt.year)}']  # first then third period of the AY
if 8 <= month <= 12:  # see if we're in the fall
    yrtrs = [str(dt.year + 1) + str(3), str(dt.year + 1) + str(5)]  # of the next year
    semesters = [f'{ay[3]} {str(dt.year)}', f'{ay[5]} {str(dt.year+1)}']  # third then fifth period of the AY
if 1 <= month <= 5:  # see if we're in the spring
    yrtrs = [str(dt.year) + str(5), str(dt.year + 1) + str(1)]  # of this and next year
    semesters = [f'{ay[5]} {str(dt.year)}', f'{ay[1]} {str(dt.year)}']  # fifth then first period of the AY

for step in range(2):
    yrtr = yrtrs[step]
    semester = semesters[step]
    print(f'Generating {semester}, using yrtr={yrtr}.')
    sheet = wb.create_sheet(index=step, title=semester)  # create a sheet, with proper title, set as active
    # url = 'https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&searchrcid=0305&searchcampusid=305&yrtr=20205&subject=ITEC&courseNumber=&courseId=&openValue=OPEN_PLUS_WAITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=&credittype=ALL&credits=&instructor=&keyword=&begindate=&site=&resultNumber=250'
    # URL of the MCTC scheduling page for ITEC, needing yrtr= to be set
    url_template = 'https://eservices.minnstate.edu/registration/search/advancedSubmit.html?campusid=305&searchrcid=0305&searchcampusid=305&yrtr=CHANGEME&subject=ITEC&courseNumber=&courseId=&openValue=OPEN_PLUS_WAITLIST&delivery=ALL&showAdvanced=&starttime=&endtime=&mntransfer=&credittype=ALL&credits=&instructor=&keyword=&begindate=&site=&resultNumber=250'
    url = url_template.replace('CHANGEME', yrtr)  # update the template URL with proper period info

    try:
        result = requests.get(url)  # try to open the URL
    except:
        print('You might not have Internet access.')  # fail gracefully
        exit(-1)
    if result.status_code != 200:  # ensure we have a legitimate response
        print('Error retrieving page.')  # fail gracefully otherwise
        exit(-2)
    # print(result.headers)  # look at the headers, just because

    src = result.content  # get the source of the page for BS
    soup = BeautifulSoup(src, 'lxml')  # use BS to get the content in a soup object
    #tbody = soup.find_all("tbody", attrs={"class":"yui-dt-data"})
    tbody = soup.find_all('tbody')  # use BS to get the particular element we want
    # two tbody's on the registration page, we want the second one
    table_rows = tbody[1].find_all('tr')  # pull all the rows <tr> from the table
    # setup a list to more easily track which index has items we want from the data
    d = {'cid': 1, 'course': 3, 'section': 4, 'title': 5, 'day': 7, 'time': 8, 'credits': 9, 'instructor': 11}
    col = 1  # start in column A

    for name in d.keys():
        sheet.cell(1, col).value = name  # print the header row in the sheet
        col += 1
    boldFont = Font(size=14, bold=True)  # make a 14 pt. bold font

    for i in range(8):  # first 8 columns; note that chr(65) is 'A', chr(66) is 'B', etc.
        sheet[f'{chr(65+i)}1'].font = boldFont  # apply the font to the header range
    sheet.row_dimensions[1].height = 14  # set row size accordingly
    sheet.freeze_panes = 'A2'  # freezes row 1
    row = 2  # skip the header row

    for tr in table_rows:  # look at each <tr> in turn
        td = tr.find_all('td')  # pull all data <td> from each row
        data = [i.text for i in td]  # make a list of all the data

        # get what we need from the data, strip whitespace
        cid = data[d['cid']].strip()
        course = data[d['course']].strip()
        section = data[d['section']].strip()
        title = data[d['title']].strip()
        day = data[d['day']].strip()
        time = data[d['time']].strip()
        course_credits = data[d['credits']].strip()
        instructor = data[d['instructor']].strip()

        # print(f'{cid:<8}{course}-{section} "{title}" on {day:^3} from {time:<14} {course_credits} credits by {instructor}')
        # set a row in the sheet with our data
        sheet.cell(row, 1, cid)
        sheet.cell(row, 2, course)
        sheet.cell(row, 3, section)
        sheet.cell(row, 4, title)
        sheet.cell(row, 5, day)
        sheet.cell(row, 6, time)
        sheet.cell(row, 7, course_credits)
        sheet.cell(row, 8, instructor)
        row += 1  # move on to the next row in the sheet

del wb['Sheet']  # get rid of the default initial sheet
wb.save(filename)  # save the workbook