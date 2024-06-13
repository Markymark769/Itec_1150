"""
Mark Porraz
4/10/2024
country list API

This program gets information and their capitals for over 200 countries

questions to ask:
Are we confident we can get the data we need?
How to get at that data?
we need to look at the structure of the JSON
List or Dictionary?

AGAIN, focus on thinking, what do we need?

"""

from openpyxl import Workbook
import requests

# TODO - make request to API to get list of country information
countries_url = 'https://country-list-1150.azurewebsites.net/api/country'
country_list_response = requests.get(countries_url).json()  # without the JSON ext we will get a code 200 error

# Descriptive variable names are good, and recommended especially in the final project
print(country_list_response)  # this is a list of country date, received in response to an API request

# TODO - create a workbook before the loop starts, so it is ready
# to have data written to it as the loop runs, and we access the country and capital
country_workbook = Workbook()
country_workbook = country_workbook.active  # select the first tab or sheet
country_workbook.title = 'Countries and Capital Cities'

row = 0

# TODO - get the country and capital city from response
# for country_data in country_list_response: - this code gets a full list of all the country data
#     print(country_data)

# for individual_country_data in country_list_response:
#     print(individual_country_dictionary)   clear that this
#     dictionary of data about one thing ( not a small list of things)


for individual_country_data in country_list_response:  # we are able to access the list by
    # using a loop for the key value pair
    country_name = individual_country_data['name']  # to be specific to get one name from this single dictionary
    print(country_name)
    # variable for capital_city here too
    # TODO - figure out how to get the capitial city
    # still inside the loop - write country name to cell
    # country_workbook.cell(1,1,country_name)  # arguments are row, col, data: this overwrites the countries until the
    # end of the list. This is a really common bug in most programs

    row = row + 1
    country_workbook.cell(1, 1, country_name)


# TODO - create a workbook

# TODO - write country and capital city to Workbook
